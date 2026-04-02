"""
Feishu Image Mapper — richData Edition (directory mode)
=========================================================
Your xlsx uses the modern "Place in cell" image format (Office 2019+),
which stores images as richData. This script reads directly from the
unzipped xlsx directory (no zip handling needed).

This script:
  1. Reads xl/metadata.xml and xl/richData/richValueRel.xml from the
     unzipped directory to map each cell -> its image filename
  2. Replaces IMAGE_PLACEHOLDER cells with =IMAGE("your_url/image.png")
  3. Clears IMAGE_PLACEHOLDER cells that have no image attached

REQUIREMENTS:
  pip install openpyxl

USAGE:
  python feishu_image_mapper.py
"""

import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

XLSX_DIR    = "l4_problematic_frames"   # unzipped xlsx folder
INPUT_FILE  = "l4_problematic_frames.xlsx"
OUTPUT_FILE = "l4_feishu_ready.xlsx"
BASE_URL    = "https://liskibruh.github.io/l4_data_sheets_table_images_mapper/l4_problematic_frames/media/"       # <- replace with your actual base URL

# ─────────────────────────────────────────────
# NAMESPACES
# ─────────────────────────────────────────────

NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_RELS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def read_xml(base_dir, relative_path):
    """Parse an XML file relative to base_dir."""
    full_path = os.path.join(base_dir, relative_path.replace("/", os.sep))
    with open(full_path, encoding="utf-8") as f:
        return ET.fromstring(f.read())


def read_rels(base_dir, xml_relative_path):
    """
    Given a path like 'xl/richData/richValueRel.xml',
    read its .rels file at 'xl/richData/_rels/richValueRel.xml.rels'.
    Returns {rId: {"target": absolute_relative_path, "type": ...}}
    """
    parts = xml_relative_path.replace("\\", "/").rsplit("/", 1)
    folder = parts[0] if len(parts) == 2 else ""
    filename = parts[-1]
    rels_relative = f"{folder}/_rels/{filename}.rels" if folder else f"_rels/{filename}.rels"

    root = read_xml(base_dir, rels_relative)
    rels = {}
    for rel in root:
        target = rel.attrib["Target"].lstrip("/")
        # Resolve relative paths like "../media/image1.png"
        if target.startswith(".."):
            base = folder
            for part in target.split("/"):
                if part == "..":
                    base = base.rsplit("/", 1)[0] if "/" in base else ""
                else:
                    base = f"{base}/{part}" if base else part
            target = base
        elif folder and not target.startswith("xl/"):
            target = f"{folder}/{target}"
        rels[rel.attrib["Id"]] = {
            "target": target,
            "type": rel.attrib.get("Type", ""),
        }
    return rels


# ─────────────────────────────────────────────
# RICHDATA PARSING
# ─────────────────────────────────────────────

def build_cell_image_map(xlsx_dir):
    """
    Returns { sheet_name: { "B3": "image1.png", ... } }
    by reading the richData XML files in the unzipped xlsx directory.
    """
    result = {}
    xl_dir = os.path.join(xlsx_dir, "xl")

    # ── 1. metadata.xml: vm-index (1-based) -> richdata index ──
    metadata_path = os.path.join(xl_dir, "metadata.xml")
    if not os.path.exists(metadata_path):
        print(f"No metadata.xml found at {metadata_path}")
        print("The file may not use richData image embedding.")
        return result

    root = ET.parse(metadata_path).getroot()
    future_meta = root.find(f"./{NS_MAIN}futureMetadata[@name='XLRICHVALUE']")
    if future_meta is None:
        print("No XLRICHVALUE metadata found — no richData images.")
        return result

    metadata_idx_to_richdata_idx = []
    for bk in future_meta:
        el = bk.find(".//*[@i]")
        if el is not None:
            metadata_idx_to_richdata_idx.append(int(el.attrib["i"]))

    print(f"metadata entries: {len(metadata_idx_to_richdata_idx)}")

    # ── 2. richValueRel.xml + its .rels: richdata index -> image filename ──
    rich_value_rel_path = "xl/richData/richValueRel.xml"
    rich_value_rels = read_rels(xlsx_dir, rich_value_rel_path)

    root = read_xml(xlsx_dir, rich_value_rel_path)
    rich_blocks = []  # ordered: index -> image filename
    for rel_el in root:
        rel_id = rel_el.attrib.get(f"{NS_RELS}id")
        if rel_id and rel_id in rich_value_rels:
            filename = os.path.basename(rich_value_rels[rel_id]["target"])
            rich_blocks.append(filename)
        else:
            rich_blocks.append(None)

    print(f"richData image blocks: {len(rich_blocks)}")

    # ── 3. workbook.xml: sheet name -> worksheet xml path ──
    workbook_rels = read_rels(xlsx_dir, "xl/workbook.xml")
    wb_root = read_xml(xlsx_dir, "xl/workbook.xml")

    sheet_name_to_xml = {}
    for sheet_el in wb_root.findall(f".//{NS_MAIN}sheet"):
        name = sheet_el.attrib["name"]
        rid  = sheet_el.attrib.get(f"{NS_RELS}id")
        if rid and rid in workbook_rels:
            sheet_name_to_xml[name] = workbook_rels[rid]["target"]

    # ── 4. Each worksheet: cells with @vm attribute have images ──
    for sheet_name, ws_xml_path in sheet_name_to_xml.items():
        cell_map = {}
        ws_root = read_xml(xlsx_dir, ws_xml_path)

        vm_cells = ws_root.findall(
            f"./{NS_MAIN}sheetData/{NS_MAIN}row/{NS_MAIN}c[@vm]"
        )
        for cell_el in vm_cells:
            vm_index = int(cell_el.attrib["vm"]) - 1  # 1-based -> 0-based
            if vm_index < len(metadata_idx_to_richdata_idx):
                rd_index = metadata_idx_to_richdata_idx[vm_index]
                if rd_index < len(rich_blocks) and rich_blocks[rd_index]:
                    coord = cell_el.attrib["r"]
                    cell_map[coord] = rich_blocks[rd_index]

        result[sheet_name] = cell_map
        print(f"  Sheet '{sheet_name}': {len(cell_map)} image cells mapped")

    return result


# ─────────────────────────────────────────────
# APPLY IMAGE URLS TO WORKBOOK
# ─────────────────────────────────────────────

def apply_image_urls(xlsx_dir, input_file, output_file, base_url):
    print(f"Parsing richData image map from: {xlsx_dir}\n")
    cell_image_map = build_cell_image_map(xlsx_dir)

    print(f"\nLoading workbook...")
    wb = load_workbook(input_file)

    total_replaced = 0
    total_cleared  = 0

    for ws in wb.worksheets:
        sheet_map = cell_image_map.get(ws.title, {})
        print(f"\nSheet '{ws.title}':")
        sheet_replaced = 0
        sheet_cleared  = 0

        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value).strip() != "#VALUE!":
                    continue

                coord    = cell.coordinate  # e.g. "B3"
                filename = sheet_map.get(coord)

                if filename:
                    url = f"{base_url}{filename}"
                    cell.value = f'=IMAGE("{url}")'
                    sheet_replaced += 1
                    total_replaced += 1
                else:
                    # Placeholder with no image — clear it
                    cell.value = ""
                    sheet_cleared += 1
                    total_cleared += 1

        print(f"  Replaced : {sheet_replaced}")
        print(f"  Cleared  : {sheet_cleared} (IMAGE_PLACEHOLDER with no image)")

    wb.save(output_file)
    print(f"\n{'='*50}")
    print(f"Done! Saved: {output_file}")
    print(f"  IMAGE() formulas written  : {total_replaced}")
    print(f"  Empty placeholders cleared: {total_cleared}")
    print(f"  Total placeholders found  : {total_replaced + total_cleared}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    # ── DIAGNOSTIC: print unique cell values to find the placeholder string ──
    wb = load_workbook(INPUT_FILE)
    unique_vals = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    unique_vals.add(repr(cell.value))
    print("Unique cell values (first 40):")
    for v in sorted(unique_vals)[:40]:
        print(" ", v)
    print()
    # ─────────────────────────────────────────────────────────────────────────

    apply_image_urls(XLSX_DIR, INPUT_FILE, OUTPUT_FILE, BASE_URL)
